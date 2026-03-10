# app.py - Dashboard AVO Carbon Group 
"""
Dashboard Flask pour visualiser les prix des métaux et les taux de change ECB.
"""

from flask import Flask, render_template, jsonify, request, send_file
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime, timedelta, date
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import calendar
import os
from decimal import Decimal

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# ===============================
# CONFIGURATION BASE DE DONNÉES
# ===============================
DB_CONFIG = {
    "user": "administrationSTS",
    "password": "St$@0987",
    "host": "avo-adb-002.postgres.database.azure.com",
    "port": "5432",
    "database": "LME_DB",
    "sslmode": "require"
}

# ==============================
# IMPORTS SUPPLÉMENTAIRES
# ==============================
try:
    from flask_mail import Mail, Message
    MAIL_AVAILABLE = True
except ImportError:
    MAIL_AVAILABLE = False
    logger.warning("flask_mail non disponible")

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    SCHEDULER_AVAILABLE = True
except ImportError:
    SCHEDULER_AVAILABLE = False
    logger.warning("apscheduler non disponible")

import atexit
import secrets

# ==============================
# CONFIGURATION EMAIL
# ==============================
app.config['MAIL_SERVER'] = 'avocarbon-com.mail.protection.outlook.com'
app.config['MAIL_PORT'] = 25
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = None
app.config['MAIL_PASSWORD'] = None
app.config['MAIL_DEFAULT_SENDER'] = 'administration.STS@avocarbon.com'

if MAIL_AVAILABLE:
    mail = Mail(app)

# ==============================
# CONFIGURATION BUDGET RATE
# ==============================
BUDGET_OWNER_EMAIL = "marjana.delija@avocarbon.com"
BUDGET_CURRENCIES = ['USD', 'CNY', 'INR', 'KRW', 'MXN', 'TND']
active_tokens = {}

# ==============================
# HELPER: SÉRIALISATION JSON ROBUSTE
# ==============================
def serialize_value(v):
    """Convertit les types PostgreSQL non-JSON-sérialisables."""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    return v

def serialize_row(row):
    """Sérialise une ligne RealDictRow complète."""
    return {k: serialize_value(v) for k, v in dict(row).items()}

# ==============================
# CONNEXION BASE DE DONNÉES
# ==============================
def get_db_connection():
    """Créer une connexion à PostgreSQL."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        logger.error(f"Erreur de connexion à la base de données: {e}")
        return None

# ==============================
# HELPER: Vérifier si une table existe
# ==============================
def table_exists(conn, table_name):
    """Vérifie si une table existe dans la DB."""
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables
                    WHERE table_schema = 'public'
                    AND table_name = %s
                )
            """, (table_name,))
            return cur.fetchone()[0]
    except Exception:
        return False

# ==============================
# MONTH HELPERS
# ==============================
def month_to_range(month_str: str):
    """
    month_str: 'YYYY-MM'
    returns (start_date, end_date) as date objects
    """
    if not month_str:
        return None, None
    try:
        y, m = map(int, month_str.split('-'))
        last_day = calendar.monthrange(y, m)[1]
        return date(y, m, 1), date(y, m, last_day)
    except Exception:
        return None, None

# ===============================
# SOURCES MÉTAUX — CONFIG
# ===============================
METALS_SOURCE_CONFIGS = {
    'brent': {
        'name': 'Brent London',
        'url_pattern': 'insee.fr',
        'url_patterns': None,
        'metal_types': ['brent_oil'],
        'currency': 'EUR',
        'unit': 'barrel',
        'format': 'year_month',
        'formula_type': 'basic_stats',
    },
    'comex': {
        'name': 'COMEX',
        'url_pattern': 'comexlive.org',
        'url_patterns': None,
        'metal_types': ['copper'],
        'currency': 'USD',
        'unit': 'lb',
        'format': 'monthly_with_conversion',
        'formula_type': 'lb_to_kg_conversion',
        'conversion_factor': 2.203,
    },
    'girm': {
        'name': 'GIRM',
        'url_pattern': 'm-lego.com',
        'url_patterns': None,
        'metal_types': ['copper'],
        'currency': 'EUR',
        'unit': 'ton',
        'format': 'yearly_columns',
        'formula_type': 'yearly_average',
    },
    'shme': {
        'name': 'SHME',
        'url_pattern': 'shmet.com',
        'url_patterns': None,
        'metal_types': ['copper', 'zinc', 'tin'],
        'currency': 'CNY',
        'unit': 'ton',
        'format': 'monthly_matrix',
        'formula_type': 'complex_alloy',
        'vat_divisor': 1.13,
        'alloy_formulas': {
            'H62':  {'cu_ratio': 0.62, 'zn_ratio': 0.38, 'base_cost': 4000, 'type': 'brass'},
            'H65':  {'cu_ratio': 0.65, 'zn_ratio': 0.35, 'base_cost': 4000, 'type': 'brass'},
            'H68':  {'cu_ratio': 0.68, 'zn_ratio': 0.32, 'base_cost': 4000, 'type': 'brass'},
            'H70':  {'cu_ratio': 0.70, 'zn_ratio': 0.30, 'base_cost': 4500, 'type': 'brass'},
            'H85':  {'cu_ratio': 0.85, 'zn_ratio': 0.15, 'base_cost': 5800, 'type': 'brass'},
            'Qsn4-0.1':   {'cu_ratio': 0.96,  'sn_ratio': 0.04,  'base_cost': 3000,  'type': 'bronze'},
            'Qsn6.5-0.1': {'cu_ratio': 0.935, 'sn_ratio': 0.065, 'base_cost': 3000,  'type': 'bronze'},
            'Qsn8-0.3':   {'cu_ratio': 0.92,  'sn_ratio': 0.08,  'base_cost': 5750,  'type': 'bronze'},
            'T2':          {'cu_ratio': 1.00,  'sn_ratio': 0.0,   'base_cost': 5500,  'type': 'tin_pure'},
        },
    },
    'lsnikko': {
        'name': 'LS NIKKO',
        'url_pattern': None,
        'url_patterns': None,
        'product_name': 'LS Nikko',
        'metal_types': ['copper'],
        'currency': 'USD',
        'unit': 'ton',
        'format': 'yearly_columns',
        'formula_type': 'yearly_average',
    },
    'silver': {
        'name': 'SILVER',
        'url_pattern': 'agosi.de',
        'url_patterns': None,
        'metal_types': ['silver'],
        'currency': 'EUR',
        'unit': 'troy_oz',
        'format': 'daily',
        'formula_type': 'basic_stats',
    },
    'lme': {
        'name': 'LME',
        'url_pattern': 'metals.dev',
        'url_patterns': ['metals.dev', 'Metal.dev API'],
        'metal_types': ['copper', 'zinc', 'tin', 'silver'],
        'currency': 'USD',
        'unit': 'ton',
        'format': 'standard',
        'formula_type': 'basic_stats',
    },
}

# ===============================
# HELPERS INTERNES MÉTAUX
# ===============================
def _build_source_filter(config, alias='mp'):
    if config.get('product_name'):
        return f"({alias}.source_product_name = %s)", [config['product_name']]
    patterns = config.get('url_patterns') or (
        [config['url_pattern']] if config.get('url_pattern') else []
    )
    if not patterns:
        return '(1=0)', []
    clauses, params = [], []
    for p in patterns:
        clauses.append(f"({alias}.source_url ILIKE %s OR {alias}.source_url = %s)")
        params.extend([f'%{p}%', p])
    return '(' + ' OR '.join(clauses) + ')', params

def _apply_date_filter(query, params, year=None, month=None,
                       start_date=None, end_date=None, date_col='price_date'):
    if year:
        query += f" AND EXTRACT(YEAR FROM {date_col}) = %s"
        params.append(int(year))
    if month:
        query += f" AND EXTRACT(MONTH FROM {date_col}) = %s"
        params.append(int(month))
    if start_date:
        try:
            query += f" AND {date_col} >= %s"
            params.append(datetime.strptime(start_date, '%Y-%m-%d').date())
        except ValueError:
            pass
    if end_date:
        try:
            query += f" AND {date_col} <= %s"
            params.append(datetime.strptime(end_date, '%Y-%m-%d').date())
        except ValueError:
            pass
    return query, params

def _serialize_metals_row(row):
    return serialize_row(row)

# ===============================
# STATISTIQUES DE CALCUL
# ===============================
def calculate_basic_stats(data):
    if isinstance(data, dict) and 'data' in data:
        prices = []
        for row in data['data']:
            for key, val in row.items():
                if key.startswith('year_') and val is not None:
                    prices.append(float(val))
    else:
        prices = [float(row.get('price', 0) or row.get('avg_price', 0) or 0)
                  for row in data
                  if row.get('price') is not None or row.get('avg_price') is not None]
    if not prices:
        return {}
    count = len(prices)
    total = sum(prices)
    average = total / count
    maximum = max(prices)
    minimum = min(prices)
    sorted_p = sorted(prices)
    median = sorted_p[count // 2] if count % 2 != 0 else \
             (sorted_p[count // 2 - 1] + sorted_p[count // 2]) / 2
    variance = sum((p - average) ** 2 for p in prices) / count
    std_dev = variance ** 0.5
    return {
        'count':         count,
        'sum':           round(total, 2),
        'average':       round(average, 2),
        'median':        round(median, 2),
        'max':           round(maximum, 2),
        'min':           round(minimum, 2),
        'range':         round(maximum - minimum, 2),
        'std_dev':       round(std_dev, 2),
        'variance':      round(variance, 2),
        'variation_pct': round(((maximum - minimum) / minimum * 100) if minimum > 0 else 0, 2),
    }

def calculate_alloy_stats(data):
    alloy_stats = {}
    for alloy_name in ['H62', 'H65', 'H68', 'H70', 'H85']:
        prices = []
        for row in data:
            if 'alloys' in row and alloy_name in row['alloys']:
                prices.append(row['alloys'][alloy_name])
        if prices:
            alloy_stats[alloy_name] = {
                'average': round(sum(prices) / len(prices), 4),
                'max':     round(max(prices), 4),
                'min':     round(min(prices), 4),
                'latest':  round(prices[0], 4),
            }
    return {
        'alloy_stats': alloy_stats,
        'base_metals': calculate_basic_stats(
            [{'price': row.get('copper_base', 0)} for row in data if row.get('copper_base')]
        ),
    }

def calculate_yearly_stats(data):
    if isinstance(data, dict) and 'years' in data:
        stats_per_year = {}
        for year in data['years']:
            key = f'year_{year}'
            prices = [row[key] for row in data['data'] if key in row and row[key] is not None]
            if prices:
                stats_per_year[str(year)] = {
                    'average': round(sum(prices) / len(prices), 2),
                    'max':     round(max(prices), 2),
                    'min':     round(min(prices), 2),
                    'count':   len(prices),
                }
        return {'yearly_stats': stats_per_year}
    return calculate_basic_stats(data)

def calculate_formulas(data, config):
    if not data:
        return {}
    ft = config.get('formula_type', 'basic_stats')
    if ft == 'complex_alloy':
        return calculate_alloy_stats(data)
    elif ft == 'yearly_average':
        return calculate_yearly_stats(data)
    else:
        return calculate_basic_stats(data)

# ===============================
# FONCTIONS RÉCUPÉRATION DONNÉES MÉTAUX
# ===============================
def get_brent_data(cursor, config, year_filter=None, month_filter=None,
                   start_date=None, end_date=None):
    src_clause, params = _build_source_filter(config)
    query = f"""
        SELECT
            EXTRACT(YEAR  FROM price_date)::INTEGER AS year,
            EXTRACT(MONTH FROM price_date)::INTEGER AS month,
            AVG(price)   AS price,
            MAX(currency) AS currency,
            COUNT(*)      AS data_points
        FROM metal_prices mp
        WHERE {src_clause}
    """
    if year_filter:
        query += " AND EXTRACT(YEAR FROM price_date) = %s"
        params.append(year_filter)
    if month_filter:
        query += " AND EXTRACT(MONTH FROM price_date) = %s"
        params.append(month_filter)
    query, params = _apply_date_filter(query, params, start_date=start_date, end_date=end_date)
    query += " GROUP BY EXTRACT(YEAR FROM price_date), EXTRACT(MONTH FROM price_date) ORDER BY year DESC, month DESC"
    cursor.execute(query, params)
    return [serialize_row(r) for r in cursor.fetchall()]

def get_shme_data(cursor, config, year_filter=None, month_filter=None,
                  start_date=None, end_date=None):
    src_clause, params = _build_source_filter(config)
    query = f"""
        SELECT
            EXTRACT(YEAR  FROM price_date)::INTEGER AS year,
            EXTRACT(MONTH FROM price_date)::INTEGER AS month,
            metal_type,
            AVG(price)    AS avg_price,
            MAX(currency) AS currency
        FROM metal_prices mp
        WHERE {src_clause}
          AND metal_type IN ('copper', 'zinc', 'tin')
    """
    if year_filter:
        query += " AND EXTRACT(YEAR FROM price_date) = %s"
        params.append(year_filter)
    if month_filter:
        query += " AND EXTRACT(MONTH FROM price_date) = %s"
        params.append(month_filter)
    query, params = _apply_date_filter(query, params, start_date=start_date, end_date=end_date)
    query += " GROUP BY EXTRACT(YEAR FROM price_date), EXTRACT(MONTH FROM price_date), metal_type ORDER BY year DESC, month DESC, metal_type"
    cursor.execute(query, params)
    base_data = cursor.fetchall()

    monthly_data = {}
    for row in base_data:
        key = f"{row['year']}-{row['month']}"
        if key not in monthly_data:
            monthly_data[key] = {
                'year': row['year'], 'month': row['month'],
                'copper': None, 'zinc': None, 'tin': None,
                'currency': row['currency']
            }
        monthly_data[key][row['metal_type']] = float(row['avg_price'])

    vat_div = config.get('vat_divisor', 1.13)
    result = []
    for key, data in sorted(monthly_data.items(), reverse=True):
        if data['copper'] is not None and data['zinc'] is not None:
            cu_raw  = data['copper']
            zn_raw  = data['zinc']
            tin_raw = data['tin']
            cu_nv   = cu_raw  / vat_div
            zn_nv   = zn_raw  / vat_div
            tin_nv  = (tin_raw / vat_div) if tin_raw is not None else None

            alloys = {}
            for grade, f in config['alloy_formulas'].items():
                alloy_type = f.get('type', 'brass')
                if alloy_type == 'brass':
                    alloys[grade] = round(
                        (cu_nv * f['cu_ratio'] * 1.05 +
                         zn_nv * f['zn_ratio'] * 1.05 +
                         f['base_cost'] / vat_div) / 1000, 4
                    )
                elif alloy_type == 'bronze' and tin_nv is not None:
                    alloys[grade] = round(
                        (cu_nv * f['cu_ratio'] * 1.05 +
                         tin_nv * f['sn_ratio'] * 1.05 +
                         f['base_cost'] / vat_div) / 1000, 4
                    )
                elif alloy_type == 'tin_pure':
                    alloys[grade] = round(
                        (cu_nv * 1.05 + f['base_cost'] / vat_div) / 1000, 4
                    )
            result.append({
                'year':        int(data['year']),
                'month':       int(data['month']),
                'copper_raw':  round(cu_raw, 2),
                'zinc_raw':    round(zn_raw, 2),
                'tin_raw':     round(tin_raw, 2) if tin_raw else None,
                'copper_base': round(cu_nv, 2),
                'zinc_base':   round(zn_nv, 2),
                'tin_base':    round(tin_nv, 2) if tin_nv else None,
                'vat_divisor': vat_div,
                'alloys':      alloys,
                'currency':    data['currency'],
            })
    return result

def get_yearly_columns_data(cursor, config, year_filter=None,
                             start_date=None, end_date=None):
    src_clause, params = _build_source_filter(config)
    query = f"""
        SELECT
            EXTRACT(YEAR  FROM price_date)::INTEGER AS year,
            EXTRACT(MONTH FROM price_date)::INTEGER AS month,
            metal_type,
            AVG(price)    AS avg_price,
            MAX(currency) AS currency
        FROM metal_prices mp
        WHERE {src_clause}
    """
    if year_filter:
        query += " AND EXTRACT(YEAR FROM price_date) = %s"
        params.append(year_filter)
    query, params = _apply_date_filter(query, params, start_date=start_date, end_date=end_date)
    query += " GROUP BY EXTRACT(YEAR FROM price_date), EXTRACT(MONTH FROM price_date), metal_type ORDER BY month, year DESC"
    cursor.execute(query, params)
    rows = cursor.fetchall()

    monthly_data, years = {}, set()
    is_girm = config.get('name') == 'GIRM'
    for row in rows:
        mo, yr = row['month'], row['year']
        years.add(yr)
        if mo not in monthly_data:
            monthly_data[mo] = {'month': mo, 'metal_type': row['metal_type']}
        val = float(row['avg_price'])
        if is_girm and val > 30:
            val = val / 100
        monthly_data[mo][f'year_{yr}'] = val

    return {
        'data':   list(monthly_data.values()),
        'years':  sorted(list(years), reverse=True),
        'format': 'yearly_columns',
    }

def get_comex_data(cursor, config, start_date=None, end_date=None,
                   year_filter=None, month_filter=None):
    src_clause, params = _build_source_filter(config)
    query = f"""
        SELECT
            EXTRACT(YEAR  FROM price_date)::INTEGER AS year,
            EXTRACT(MONTH FROM price_date)::INTEGER AS month,
            AVG(price)    AS avg_price_lb,
            MAX(currency) AS currency
        FROM metal_prices mp
        WHERE {src_clause}
    """
    if year_filter:
        query += " AND EXTRACT(YEAR FROM price_date) = %s"
        params.append(year_filter)
    if month_filter:
        query += " AND EXTRACT(MONTH FROM price_date) = %s"
        params.append(month_filter)
    if start_date:
        query += " AND price_date >= %s"
        params.append(start_date)
    if end_date:
        query += " AND price_date <= %s"
        params.append(end_date)
    else:
        query += " AND price_date >= CURRENT_DATE - INTERVAL '2 years'"
    query += " GROUP BY EXTRACT(YEAR FROM price_date), EXTRACT(MONTH FROM price_date) ORDER BY year DESC, month DESC"
    cursor.execute(query, params)
    rows = cursor.fetchall()

    factor = config['conversion_factor']
    result = []
    for row in rows:
        price_lb = float(row['avg_price_lb'])
        result.append({
            'year':               int(row['year']),
            'month':              int(row['month']),
            'price_lb':           price_lb,
            'price_kg_usd':       round(price_lb * factor, 4),
            'conversion_formula': f"price_lb * {factor}",
            'currency':           row['currency'],
        })
    return result

def get_standard_data(cursor, config, start_date=None, end_date=None, metal_type=None):
    src_clause, params = _build_source_filter(config)
    query = f"""
        SELECT price_date, metal_type, price, currency, unit, source_url
        FROM metal_prices mp
        WHERE {src_clause}
    """
    if start_date:
        query += " AND price_date >= %s"
        params.append(start_date)
    if end_date:
        query += " AND price_date <= %s"
        params.append(end_date)
    else:
        query += " AND price_date >= CURRENT_DATE - INTERVAL '1 year'"
    if metal_type and metal_type != 'all':
        query += " AND metal_type = %s"
        params.append(metal_type)
    query += " ORDER BY price_date DESC, metal_type"
    cursor.execute(query, params)
    return [_serialize_metals_row(r) for r in cursor.fetchall()]

# ===============================
# FONCTIONS GÉNÉRALES
# ===============================
def get_latest_prices():
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                WITH latest_prices AS (
                    SELECT id, metal_type,
                        ROW_NUMBER() OVER (PARTITION BY metal_type ORDER BY price_date DESC, created_at DESC) AS rn
                    FROM metal_prices
                )
                SELECT p.* FROM metal_prices p JOIN latest_prices lp ON p.id = lp.id WHERE lp.rn = 1
                ORDER BY p.metal_type;
            """)
            return cur.fetchall()
    except Exception as e:
        logger.error(f"Erreur get_latest_prices: {e}")
        return []
    finally:
        conn.close()

def get_price_history(days=None, metal_type=None, start_date=None, end_date=None, month=None):
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            query = """
                SELECT metal_type, price, currency, unit, source_url, price_date, created_at
                FROM metal_prices WHERE 1=1
            """
            params = []
            if month and not start_date and not end_date:
                sd, ed = month_to_range(month)
                if sd and ed:
                    query += " AND price_date >= %s AND price_date <= %s"
                    params.extend([sd, ed])
            elif start_date or end_date:
                if start_date:
                    try:
                        query += " AND price_date >= %s"
                        params.append(datetime.strptime(start_date, '%Y-%m-%d').date())
                    except ValueError:
                        pass
                if end_date:
                    try:
                        query += " AND price_date <= %s"
                        params.append(datetime.strptime(end_date, '%Y-%m-%d').date())
                    except ValueError:
                        pass
            elif days:
                query += " AND price_date >= %s"
                params.append((datetime.now() - timedelta(days=int(days))).date())
            if metal_type and metal_type.lower() != 'all':
                query += " AND metal_type = %s"
                params.append(metal_type)
            query += " ORDER BY metal_type, price_date DESC, created_at DESC"
            cur.execute(query, params)
            return cur.fetchall()
    except Exception as e:
        logger.error(f"Erreur get_price_history: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return []
    finally:
        conn.close()

def get_statistics():
    """
    FIX: Sérialise les Decimal → float pour éviter les crashes JSON.
    """
    conn = get_db_connection()
    if not conn:
        return {'total_records': 0, 'total_metals': 0, 'variations': []}
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT COUNT(*) AS total_records, COUNT(DISTINCT metal_type) AS total_metals
                FROM metal_prices;
            """)
            summary = cur.fetchone()
            cur.execute("""
                WITH R AS (
                    SELECT metal_type, price, currency, price_date, created_at,
                        ROW_NUMBER() OVER (
                            PARTITION BY metal_type
                            ORDER BY price_date DESC, created_at DESC
                        ) AS rn
                    FROM metal_prices
                )
                SELECT
                    l.metal_type,
                    l.price          AS current_price,
                    l.currency,
                    p.price          AS previous_price,
                    CASE
                        WHEN p.price IS NOT NULL AND p.price != 0
                        THEN ((l.price - p.price) / p.price) * 100
                        ELSE NULL
                    END AS variation_percent
                FROM (SELECT metal_type, price, currency FROM R WHERE rn=1) l
                LEFT JOIN (SELECT metal_type, price FROM R WHERE rn=2) p
                    ON l.metal_type = p.metal_type
                ORDER BY l.metal_type;
            """)
            variations_raw = cur.fetchall()
            # ✅ FIX: Sérialiser chaque ligne pour convertir les Decimal
            variations = [serialize_row(v) for v in variations_raw]

            return {
                'total_records': int(summary['total_records']),
                'total_metals':  int(summary['total_metals']),
                'variations':    variations,
            }
    except Exception as e:
        logger.error(f"Erreur get_statistics: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {'total_records': 0, 'total_metals': 0, 'variations': []}
    finally:
        conn.close()

# ===============================
# API DYNAMIQUE: METAL TYPES
# ===============================
def get_all_metal_types():
    """Récupère la liste de tous les types de métaux distincts en DB."""
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT metal_type FROM metal_prices
                WHERE metal_type IS NOT NULL
                ORDER BY metal_type;
            """)
            return [row[0] for row in cur.fetchall()]
    except Exception as e:
        logger.error(f"Erreur get_all_metal_types: {e}")
        return []
    finally:
        conn.close()

def get_all_sources():
    """Récupère la liste de toutes les source_url distinctes en DB."""
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT source_url FROM metal_prices
                WHERE source_url IS NOT NULL
                ORDER BY source_url;
            """)
            return [row[0] for row in cur.fetchall()]
    except Exception as e:
        logger.error(f"Erreur get_all_sources: {e}")
        return []
    finally:
        conn.close()

def get_all_fx_currencies():
    """Récupère la liste de toutes les devises ECB en DB."""
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT quote_currency FROM ecb_exchange_rates
                WHERE quote_currency IS NOT NULL
                ORDER BY quote_currency;
            """)
            return [row[0] for row in cur.fetchall()]
    except Exception as e:
        logger.error(f"Erreur get_all_fx_currencies: {e}")
        return []
    finally:
        conn.close()

def get_price_date_range():
    """Retourne min/max price_date pour initialiser les datepickers."""
    conn = get_db_connection()
    if not conn:
        return {}
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT MIN(price_date)::text, MAX(price_date)::text FROM metal_prices;")
            row = cur.fetchone()
            return {'min_date': row[0], 'max_date': row[1]}
    except Exception as e:
        logger.error(f"Erreur get_price_date_range: {e}")
        return {}
    finally:
        conn.close()

def get_fx_date_range():
    """Retourne min/max ref_date pour les filtres FX."""
    conn = get_db_connection()
    if not conn:
        return {}
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT MIN(ref_date)::text, MAX(ref_date)::text FROM ecb_exchange_rates;")
            row = cur.fetchone()
            return {'min_date': row[0], 'max_date': row[1]}
    except Exception as e:
        logger.error(f"Erreur get_fx_date_range: {e}")
        return {}
    finally:
        conn.close()

# ===============================
# ECB / FX FUNCTIONS
# ===============================
def get_ecb_rates(start_date=None, end_date=None, quote_currency=None, month=None):
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            query = """
                SELECT
                    ref_date,
                    base_currency,
                    quote_currency,
                    rate,
                    source_url,
                    metadata
                FROM ecb_exchange_rates
                WHERE 1=1
            """
            params = []
            if month:
                sd, ed = month_to_range(month)
                if sd and ed:
                    query += " AND ref_date >= %s AND ref_date <= %s"
                    params.extend([sd, ed])
            else:
                if start_date:
                    try:
                        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
                        query += " AND ref_date >= %s"
                        params.append(sd)
                    except ValueError:
                        pass
                if end_date:
                    try:
                        ed = datetime.strptime(end_date, "%Y-%m-%d").date()
                        query += " AND ref_date <= %s"
                        params.append(ed)
                    except ValueError:
                        pass
                if not start_date and not end_date:
                    query += " AND ref_date >= %s"
                    params.append(datetime.now().date() - timedelta(days=365))

            if quote_currency and quote_currency.lower() != 'all':
                query += " AND quote_currency = %s"
                params.append(quote_currency.upper())

            query += " ORDER BY ref_date DESC, quote_currency ASC"
            cur.execute(query, params)
            return cur.fetchall()
    except Exception as e:
        logger.error(f"Erreur get_ecb_rates: {e}")
        return []
    finally:
        conn.close()

def get_florent_report_data(year, month):
    """
    Rapport Florent avec Budget Rate.
    FIX: Correction de l'ordre des paramètres SQL.
    """
    conn = get_db_connection()
    if not conn:
        return []

    # Vérifier si la table fx_budget_rates existe
    has_budget_table = table_exists(conn, 'fx_budget_rates')

    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if has_budget_table:
                budget_join = "LEFT JOIN fx_budget_rates br ON md.quote_currency = br.currency AND br.year = %(year_budget)s"
                budget_select = ", br.budget_rate"
            else:
                budget_join = ""
                budget_select = ", NULL::numeric AS budget_rate"

            query = f"""
            WITH MonthlyData AS (
                SELECT
                    quote_currency,
                    rate,
                    ref_date,
                    LAST_VALUE(rate) OVER (
                        PARTITION BY quote_currency, DATE_TRUNC('month', ref_date)
                        ORDER BY ref_date
                        RANGE BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING
                    ) AS month_closing
                FROM ecb_exchange_rates
                WHERE EXTRACT(YEAR FROM ref_date) = %(year)s
                   OR (
                        EXTRACT(YEAR FROM ref_date)  = %(year)s - 1
                        AND %(month)s = 1
                        AND EXTRACT(MONTH FROM ref_date) = 12
                   )
            ),
            MonthlyPeriodRates AS (
                SELECT
                    quote_currency,
                    EXTRACT(MONTH FROM ref_date) AS m,
                    AVG(rate) AS period_rate
                FROM MonthlyData
                GROUP BY quote_currency, EXTRACT(MONTH FROM ref_date)
            )
            SELECT
                md.quote_currency,
                MAX(
                    CASE WHEN EXTRACT(MONTH FROM md.ref_date) = %(month)s
                    THEN md.month_closing END
                ) AS closing_rate,
                MAX(
                    CASE
                        WHEN %(month)s = 1 AND EXTRACT(MONTH FROM md.ref_date) = 12
                        THEN md.month_closing
                        WHEN %(month)s > 1 AND EXTRACT(MONTH FROM md.ref_date) = %(month)s - 1
                        THEN md.month_closing
                    END
                ) AS period_rate,
                CASE
                    WHEN %(month)s = 1 THEN (
                        SELECT mpr.period_rate FROM MonthlyPeriodRates mpr
                        WHERE mpr.quote_currency = md.quote_currency AND mpr.m = 1
                    )
                    ELSE (
                        SELECT AVG(mpr.period_rate) FROM MonthlyPeriodRates mpr
                        WHERE mpr.quote_currency = md.quote_currency AND mpr.m <= %(month)s
                    )
                END AS ytd_average
                {budget_select}
            FROM MonthlyData md
            {budget_join}
            GROUP BY md.quote_currency{', br.budget_rate' if has_budget_table else ''}
            ORDER BY md.quote_currency;
            """

            params = {'year': year, 'month': month}
            if has_budget_table:
                params['year_budget'] = year

            cur.execute(query, params)
            rows = cur.fetchall()
            return [serialize_row(r) for r in rows]

    except Exception as e:
        logger.error(f"Erreur get_florent_report_data: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return []
    finally:
        conn.close()

def get_monthly_fx_summary(year=None, month=None, quote_currency=None):
    """
    Résumé mensuel FX.
    FIX: Guard sur la table fx_budget_rates (peut ne pas exister).
    """
    conn = get_db_connection()
    if not conn:
        return []
    try:
        if not year or not month:
            today = datetime.now()
            year = today.year
            month = today.month
        else:
            year = int(year)
            month = int(month)

        has_budget_table = table_exists(conn, 'fx_budget_rates')

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if has_budget_table:
                budget_join   = "LEFT JOIN fx_budget_rates br ON mc.quote_currency = br.currency AND br.year = %s"
                budget_select = ", br.budget_rate"
                budget_group  = ", br.budget_rate"
            else:
                budget_join   = ""
                budget_select = ", NULL::numeric AS budget_rate"
                budget_group  = ""

            query = f"""
            WITH MonthlyClosing AS (
                SELECT DISTINCT ON (quote_currency)
                    quote_currency,
                    rate     AS closing_rate,
                    ref_date AS closing_date
                FROM ecb_exchange_rates
                WHERE EXTRACT(YEAR  FROM ref_date) = %s
                  AND EXTRACT(MONTH FROM ref_date) = %s
                ORDER BY quote_currency, ref_date DESC
            ),
            PreviousMonthClosing AS (
                SELECT DISTINCT ON (quote_currency)
                    quote_currency,
                    rate     AS period_rate,
                    ref_date AS period_date
                FROM ecb_exchange_rates
                WHERE (
                    (EXTRACT(YEAR  FROM ref_date) = %s AND EXTRACT(MONTH FROM ref_date) = %s - 1)
                    OR
                    (EXTRACT(YEAR  FROM ref_date) = %s - 1 AND EXTRACT(MONTH FROM ref_date) = 12 AND %s = 1)
                )
                ORDER BY quote_currency, ref_date DESC
            ),
            YTDAverage AS (
                SELECT
                    quote_currency,
                    AVG(rate) AS ytd_average
                FROM ecb_exchange_rates
                WHERE EXTRACT(YEAR  FROM ref_date) = %s
                  AND EXTRACT(MONTH FROM ref_date) <= %s
                GROUP BY quote_currency
            )
            SELECT
                mc.quote_currency,
                mc.closing_rate,
                mc.closing_date,
                pmc.period_rate,
                pmc.period_date,
                ytd.ytd_average
                {budget_select}
            FROM MonthlyClosing mc
            LEFT JOIN PreviousMonthClosing pmc ON mc.quote_currency = pmc.quote_currency
            LEFT JOIN YTDAverage ytd           ON mc.quote_currency = ytd.quote_currency
            {budget_join}
            WHERE 1=1
            """

            # Paramètres dans le bon ordre
            params = [
                year, month,        # MonthlyClosing
                year, month,        # PreviousMonthClosing (mois-1)
                year, month,        # PreviousMonthClosing (décembre an-1 si janvier)
                year, month,        # YTDAverage
            ]
            if has_budget_table:
                params.append(year)  # budget join year

            if quote_currency and quote_currency.lower() != 'all':
                query += " AND mc.quote_currency = %s"
                params.append(quote_currency.upper())

            query += f" ORDER BY mc.quote_currency"

            cur.execute(query, params)
            rows = cur.fetchall()
            return [serialize_row(r) for r in rows]

    except Exception as e:
        logger.error(f"Erreur get_monthly_fx_summary: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return []
    finally:
        conn.close()

def get_sync_logs(limit=10):
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, sync_type, status, metals_updated, error_message,
                       duration_seconds, created_at
                FROM sync_logs ORDER BY created_at DESC LIMIT %s;
            """, (limit,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"Erreur get_sync_logs: {e}")
        return []
    finally:
        conn.close()

# ==============================
# BUDGET RATE — EMAIL + FORMULAIRE
# ==============================
def generate_secure_token(year):
    token = secrets.token_urlsafe(32)
    active_tokens[token] = {
        'year': year,
        'created_at': datetime.now(),
        'used': False
    }
    return token

def get_email_html_template(year, form_url, token):
    return f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8"><title>Budget FX Rates {year}</title></head>
<body style="font-family:Segoe UI,Arial,sans-serif;background:#f5f7fa;padding:20px;">
<div style="max-width:700px;margin:0 auto;background:white;border-radius:12px;overflow:hidden;">
  <div style="background:linear-gradient(135deg,#0066b2,#004d8c);color:white;padding:30px;text-align:center;">
    <h1 style="margin:0;">Budget FX Rates {year}</h1>
    <p>Formulaire de saisie annuel</p>
  </div>
  <div style="padding:30px;">
    <p>Il est temps de définir les <strong>Budget Rates</strong> pour <strong>{year}</strong>.</p>
    <p style="text-align:center;margin:30px 0;">
      <a href="{form_url}" style="background:#27ae60;color:white;padding:15px 30px;
         text-decoration:none;border-radius:8px;font-weight:600;">
        ✏️ Remplir le formulaire Budget {year}
      </a>
    </p>
    <p style="font-size:12px;color:#7f8c8d;">Token: <code>{token}</code></p>
  </div>
</div>
</body></html>"""

def get_form_html_template(year, token, existing_rates=None):
    existing_rates = existing_rates or {}
    currency_inputs = ""
    for curr in BUDGET_CURRENCIES:
        value = existing_rates.get(curr, "")
        currency_inputs += f"""
        <div style="display:flex;flex-direction:column;gap:5px;">
          <label style="font-weight:600;color:#2c3e50;">{curr}</label>
          <input type="number" name="{curr}" id="{curr}" step="0.0001"
            placeholder="1.1000" value="{value}" required
            style="padding:10px;border:2px solid #ecf0f1;border-radius:6px;font-size:14px;">
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8"><title>Budget FX Rates {year}</title></head>
<body style="font-family:Segoe UI,Arial,sans-serif;background:linear-gradient(135deg,#667eea,#764ba2);min-height:100vh;padding:20px;">
<div style="max-width:700px;margin:0 auto;background:white;border-radius:15px;overflow:hidden;">
  <div style="background:linear-gradient(135deg,#0066b2,#004d8c);color:white;padding:30px;text-align:center;">
    <h1>📊 Budget FX Rates {year}</h1>
  </div>
  <div style="padding:30px;">
    <div id="message"></div>
    <form id="budgetForm">
      <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:15px;margin:20px 0;">
        {currency_inputs}
      </div>
      <button type="submit" id="submitBtn"
        style="width:100%;background:linear-gradient(135deg,#27ae60,#229954);color:white;
               border:none;padding:15px;font-size:16px;font-weight:600;border-radius:8px;cursor:pointer;">
        ✅ Enregistrer Budget Rates {year}
      </button>
    </form>
  </div>
</div>
<script>
document.getElementById('budgetForm').addEventListener('submit', async (e) => {{
  e.preventDefault();
  const btn = document.getElementById('submitBtn');
  btn.disabled = true; btn.textContent = 'Enregistrement...';
  const rates = {{}};
  document.querySelectorAll('input[name]').forEach(i => {{ rates[i.name] = parseFloat(i.value); }});
  try {{
    const r = await fetch('/api/submit-budget-rates', {{
      method:'POST', headers:{{'Content-Type':'application/json'}},
      body: JSON.stringify({{token:'{token}', year:{year}, rates}})
    }});
    const j = await r.json();
    const msg = document.getElementById('message');
    if (j.status === 'success') {{
      msg.innerHTML = '<div style="background:#d4edda;border-left:4px solid #27ae60;padding:15px;border-radius:5px;margin-bottom:15px;"><strong>✅ Succès!</strong> ' + j.message + '</div>';
      document.querySelectorAll('input').forEach(i => i.disabled = true);
      btn.textContent = '✅ Enregistré';
    }} else {{
      throw new Error(j.message);
    }}
  }} catch(err) {{
    document.getElementById('message').innerHTML = '<div style="background:#f8d7da;border-left:4px solid #e74c3c;padding:15px;border-radius:5px;margin-bottom:15px;"><strong>❌ Erreur:</strong> ' + err.message + '</div>';
    btn.disabled = false; btn.textContent = '✅ Enregistrer Budget Rates {year}';
  }}
}});
</script>
</body></html>"""

def send_budget_rate_email(year, recipient_email=BUDGET_OWNER_EMAIL, test_mode=False):
    try:
        token = generate_secure_token(year)
        if test_mode:
            form_url = f"http://localhost:5000/budget-form/{token}"
            logger.info(f"🧪 TEST - URL formulaire: {form_url}")
            return True
        form_url = f"https://avo-exmetrics.azurewebsites.net/budget-form/{token}"
        email_html = get_email_html_template(year, form_url, token)
        if MAIL_AVAILABLE:
            msg = Message(
                subject=f"[ACTION REQUISE] Budget FX Rates {year}",
                recipients=[recipient_email],
                html=email_html
            )
            mail.send(msg)
        logger.info(f"✅ Email Budget Rate {year} envoyé à {recipient_email}")
        return True
    except Exception as e:
        logger.error(f"❌ Erreur envoi email Budget Rate: {e}")
        return False

# ==============================
# ROUTES BUDGET
# ==============================
@app.route('/budget-form/<token>')
def budget_form(token):
    if token not in active_tokens:
        return "<h1>❌ Lien invalide ou expiré</h1>", 403
    token_data = active_tokens[token]
    if token_data['used']:
        return "<h1>✅ Formulaire déjà soumis</h1>", 200
    if (datetime.now() - token_data['created_at']).days > 30:
        return "<h1>⏰ Lien expiré</h1>", 410
    year = token_data['year']
    existing_rates = {}
    conn = get_db_connection()
    if conn and table_exists(conn, 'fx_budget_rates'):
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT currency, budget_rate FROM fx_budget_rates WHERE year = %s", (year,))
                for row in cur.fetchall():
                    existing_rates[row['currency']] = float(row['budget_rate'])
        except Exception as e:
            logger.error(f"Erreur récupération taux existants: {e}")
        finally:
            conn.close()
    return get_form_html_template(year, token, existing_rates)

@app.route('/api/submit-budget-rates', methods=['POST'])
def submit_budget_rates():
    try:
        data = request.json
        token = data.get('token')
        year  = data.get('year')
        rates = data.get('rates', {})
        if token not in active_tokens:
            return jsonify({'status': 'error', 'message': 'Token invalide'}), 403
        token_data = active_tokens[token]
        if token_data['used']:
            return jsonify({'status': 'error', 'message': 'Déjà soumis'}), 400
        if (datetime.now() - token_data['created_at']).days > 30:
            return jsonify({'status': 'error', 'message': 'Lien expiré'}), 410
        if year != token_data['year']:
            return jsonify({'status': 'error', 'message': 'Année invalide'}), 400
        if not rates:
            return jsonify({'status': 'error', 'message': 'Aucun taux'}), 400
        conn = get_db_connection()
        if not conn:
            return jsonify({'status': 'error', 'message': 'Erreur DB'}), 500
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM fx_budget_rates WHERE year = %s", (year,))
                for currency, rate in rates.items():
                    cur.execute("""
                        INSERT INTO fx_budget_rates (year, currency, budget_rate, updated_at)
                        VALUES (%s, %s, %s, NOW())
                    """, (year, currency, rate))
                conn.commit()
            active_tokens[token]['used'] = True
            return jsonify({'status': 'success', 'message': f'Budget Rates {year} enregistrés ({len(rates)} devises)'})
        except Exception as e:
            conn.rollback()
            return jsonify({'status': 'error', 'message': str(e)}), 500
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==============================
# CRON SCHEDULER
# ==============================
if SCHEDULER_AVAILABLE:
    def scheduled_budget_email_job():
        with app.app_context():
            try:
                next_year = datetime.now().year + 1
                send_budget_rate_email(year=next_year, recipient_email=BUDGET_OWNER_EMAIL, test_mode=False)
            except Exception as e:
                logger.error(f"Erreur cron Budget Rate: {e}")

    scheduler = BackgroundScheduler()
    scheduler.add_job(
        func=scheduled_budget_email_job,
        trigger=CronTrigger(month=11, day=1, hour=9, minute=0),
        id="budget_rate_annual_email",
        replace_existing=True
    )
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        scheduler.start()
        logger.info("✅ Scheduler démarré")
    atexit.register(lambda: scheduler.shutdown())

# ==============================
# ROUTES TEST
# ==============================
@app.route('/test-budget-email')
def test_budget_email():
    year  = request.args.get('year', type=int, default=datetime.now().year + 1)
    email = request.args.get('email', default=BUDGET_OWNER_EMAIL)
    success = send_budget_rate_email(year=year, recipient_email=email, test_mode=True)
    if success and active_tokens:
        latest_token = list(active_tokens.keys())[-1]
        form_url = f"http://localhost:5000/budget-form/{latest_token}"
        return f"<h1>✅ Test OK</h1><p>Token: {latest_token}</p><a href='{form_url}'>Ouvrir formulaire</a>"
    return "<h1>❌ Erreur</h1>", 500

# ===============================
# ROUTES FLASK PRINCIPALES
# ===============================
@app.route('/')
def landing_page():
    return render_template('landing.html')

@app.route('/dashboard')
def dashboard():
    return render_template('index.html')

@app.route('/health')
def health_check():
    try:
        conn = get_db_connection()
        if conn:
            conn.close()
            return jsonify({'status': 'ok', 'db': 'connected'}), 200
        return jsonify({'status': 'error', 'db': 'disconnected'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ──────────────────────────────────────────
# ✅ NOUVELLES ROUTES: Filtres dynamiques
# ──────────────────────────────────────────

@app.route('/api/metals/metal-types')
def api_metal_types():
    """
    Retourne la liste des types de métaux disponibles en DB.
    Utilisé pour peupler dynamiquement le dropdown côté frontend.
    """
    metal_types = get_all_metal_types()
    return jsonify({'status': 'success', 'data': metal_types})

@app.route('/api/metals/sources')
def api_metal_sources():
    """
    Retourne la liste des sources disponibles en DB.
    Utilisé pour peupler dynamiquement le dropdown source.
    """
    sources = get_all_sources()
    return jsonify({'status': 'success', 'data': sources})

@app.route('/api/metals/date-range')
def api_metals_date_range():
    """
    Retourne min/max price_date depuis la DB.
    Utilisé pour initialiser les datepickers avec les vraies bornes.
    """
    dr = get_price_date_range()
    return jsonify({'status': 'success', 'data': dr})

@app.route('/api/fx/currencies')
def api_fx_currencies():
    """
    Retourne la liste de toutes les devises ECB disponibles en DB.
    Utilisé pour peupler dynamiquement le dropdown FX.
    """
    currencies = get_all_fx_currencies()
    return jsonify({'status': 'success', 'data': currencies})

@app.route('/api/fx/date-range')
def api_fx_date_range():
    """
    Retourne min/max ref_date depuis la DB.
    Utilisé pour initialiser les datepickers FX avec les vraies bornes.
    """
    dr = get_fx_date_range()
    return jsonify({'status': 'success', 'data': dr})

# ──────────────────────────────────────────
# API MÉTAUX
# ──────────────────────────────────────────

@app.route('/api/prices/latest')
def api_latest_prices():
    prices = get_latest_prices()
    return jsonify({'status': 'success', 'data': [serialize_row(p) for p in prices]})

@app.route('/api/prices/history')
def api_price_history():
    days       = request.args.get('days', type=int)
    metal_type = request.args.get('metal_type')
    start_date = request.args.get('start_date')
    end_date   = request.args.get('end_date')
    month      = request.args.get('month')

    if metal_type and metal_type.lower() == 'all':
        metal_type = None

    if month:
        sd, ed = month_to_range(month)
        if sd and ed:
            ms, me = sd, ed
            if start_date:
                try:
                    ms = max(ms, datetime.strptime(start_date, "%Y-%m-%d").date())
                except ValueError:
                    pass
            if end_date:
                try:
                    me = min(me, datetime.strptime(end_date, "%Y-%m-%d").date())
                except ValueError:
                    pass
            start_date, end_date, days = ms.isoformat(), me.isoformat(), None

    history = get_price_history(days, metal_type, start_date, end_date)

    return jsonify({
        'status': 'success',
        'data': [serialize_row(i) for i in history]
    })

@app.route('/api/statistics')
def api_statistics():
    """FIX: Retourne des données JSON-sérialisables (Decimal → float)."""
    stats = get_statistics()
    return jsonify({'status': 'success', 'data': stats})

@app.route('/api/sync/logs')
def api_sync_logs():
    logs = get_sync_logs()
    return jsonify({'status': 'success', 'data': [serialize_row(l) for l in logs]})

# ──────────────────────────────────────────
# WORKBOOK ROUTES
# ──────────────────────────────────────────

def get_bme_data(cursor, year_filter=None, month_filter=None):
    """
    BME: Exchange rate matrix from ecb_exchange_rates.
    FIX: Utilise ecb_exchange_rates (et non ecb_rates).
    """
    yr = int(year_filter) if year_filter else datetime.now().year
    params = [yr]
    query = """
        SELECT quote_currency,
               EXTRACT(MONTH FROM ref_date)::INTEGER AS month,
               AVG(rate) AS avg_rate
        FROM ecb_exchange_rates
        WHERE EXTRACT(YEAR FROM ref_date) = %s
    """
    if month_filter:
        query += " AND EXTRACT(MONTH FROM ref_date) = %s"
        params.append(int(month_filter))
    query += " GROUP BY quote_currency, EXTRACT(MONTH FROM ref_date) ORDER BY quote_currency, month"
    cursor.execute(query, params)
    rows = cursor.fetchall()

    pivot = {}
    currencies = []
    for row in rows:
        ccy  = row['quote_currency']
        mo   = row['month']
        rate = float(row['avg_rate'])
        if ccy not in pivot:
            pivot[ccy] = {}
            currencies.append(ccy)
        pivot[ccy][mo] = round(rate, 6)

    data = []
    for ccy in currencies:
        entry = {'currency': ccy, 'pair': f'EUR/{ccy}'}
        for m in range(1, 13):
            entry[f'month_{m}'] = pivot[ccy].get(m)
        data.append(entry)

    return {'data': data, 'currencies': currencies,
            'months': list(range(1, 13)), 'year': yr}

@app.route('/metals-workbook')
def metals_workbook():
    return render_template('metals_workbook.html')

@app.route('/api/metals/sheets')
def api_metals_sheets():
    """Résumé de toutes les sources pour les badges workbook."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'status': 'error', 'message': 'DB connection failed'}), 500
    try:
        result = {}
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            for sheet_id, config in METALS_SOURCE_CONFIGS.items():
                fmt = config.get('format', 'standard')
                if fmt == 'exchange_matrix':
                    # FIX: utilise ecb_exchange_rates
                    cur.execute("""
                        SELECT COUNT(*) AS cnt,
                               MAX(ref_date) AS last_date,
                               MIN(ref_date) AS first_date
                        FROM ecb_exchange_rates
                    """)
                    row = cur.fetchone()
                    result[sheet_id] = {
                        'name':       config['name'],
                        'count':      int(row['cnt']) if row else 0,
                        'last_date':  row['last_date'].isoformat() if row and row['last_date'] else None,
                        'first_date': row['first_date'].isoformat() if row and row['first_date'] else None,
                        'format':     fmt
                    }
                    continue

                if config.get('product_name'):
                    cur.execute("""
                        SELECT COUNT(*) AS cnt,
                               MAX(price_date) AS last_date,
                               MIN(price_date) AS first_date
                        FROM metal_prices WHERE source_product_name = %s
                    """, [config['product_name']])
                    row = cur.fetchone()
                    result[sheet_id] = {
                        'name':       config['name'],
                        'count':      int(row['cnt']) if row else 0,
                        'last_date':  row['last_date'].isoformat() if row and row['last_date'] else None,
                        'first_date': row['first_date'].isoformat() if row and row['first_date'] else None,
                        'format':     fmt
                    }
                    continue

                patterns = config.get('url_patterns') or (
                    [config['url_pattern']] if config.get('url_pattern') else []
                )
                if not patterns:
                    result[sheet_id] = {'name': config['name'], 'count': 0,
                                        'last_date': None, 'first_date': None, 'format': fmt}
                    continue

                clauses, params = [], []
                for p in patterns:
                    clauses.append('(source_url ILIKE %s OR source_url = %s)')
                    params.extend([f'%{p}%', p])
                where = ' OR '.join(clauses)
                cur.execute(f"""
                    SELECT COUNT(*) AS cnt,
                           MAX(price_date) AS last_date,
                           MIN(price_date) AS first_date
                    FROM metal_prices WHERE {where}
                """, params)
                row = cur.fetchone()
                result[sheet_id] = {
                    'name':       config['name'],
                    'count':      int(row['cnt']) if row else 0,
                    'last_date':  row['last_date'].isoformat() if row and row['last_date'] else None,
                    'first_date': row['first_date'].isoformat() if row and row['first_date'] else None,
                    'format':     fmt
                }
        return jsonify({'status': 'success', 'sheets': result})
    except Exception as e:
        logger.error(f"Erreur api_metals_sheets: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/metals/sheet/<sheet_id>')
def api_get_sheet_data(sheet_id):
    # SUMMARY tab is under enhancement — return 200 instead of 400
    if sheet_id == 'summary':
        return jsonify({'status': 'enhancement', 'sheet_id': 'summary', 'message': 'En cours — Phase Enhancement', 'data': []}), 200

    if sheet_id not in METALS_SOURCE_CONFIGS:
        return jsonify({'status': 'error',
                        'message': f"Sheet '{sheet_id}' invalide."}), 400

    config       = METALS_SOURCE_CONFIGS[sheet_id]
    year_filter  = request.args.get('year',       type=int)
    month_filter = request.args.get('month',      type=int)
    start_date   = request.args.get('start_date')
    end_date     = request.args.get('end_date')
    metal_type   = request.args.get('metal_type')

    conn = get_db_connection()
    if not conn:
        return jsonify({'status': 'error', 'message': 'DB connection failed'}), 500

    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            fmt = config.get('format', 'standard')

            if fmt == 'exchange_matrix':
                bme = get_bme_data(cur, year_filter, month_filter)
                return jsonify({
                    'status': 'success', 'sheet_id': sheet_id,
                    'sheet_name': config['name'],
                    'data': bme.get('data', []),
                    'currencies': bme.get('currencies', []),
                    'months': bme.get('months', list(range(1, 13))),
                    'year': bme.get('year'),
                    'formulas': {},
                    'config': {'format': fmt, 'formula_type': config.get('formula_type')}
                })
            elif fmt == 'year_month':
                data = get_brent_data(cur, config, year_filter, month_filter, start_date, end_date)
            elif fmt == 'monthly_matrix':
                data = get_shme_data(cur, config, year_filter, month_filter, start_date, end_date)
            elif fmt == 'yearly_columns':
                data = get_yearly_columns_data(cur, config, year_filter, start_date, end_date)
            elif fmt == 'monthly_with_conversion':
                data = get_comex_data(cur, config, start_date, end_date, year_filter, month_filter)
            else:
                data = get_standard_data(cur, config, start_date, end_date, metal_type)

            formulas = calculate_formulas(data, config)
            return jsonify({
                'status':     'success',
                'sheet_id':   sheet_id,
                'sheet_name': config['name'],
                'data':       data,
                'formulas':   formulas,
                'config': {
                    'format':       config.get('format'),
                    'formula_type': config.get('formula_type'),
                },
            })
    except Exception as e:
        logger.error(f"Erreur api_get_sheet_data [{sheet_id}]: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/metals/export/<sheet_id>')
def export_sheet_excel(sheet_id):
    if sheet_id not in METALS_SOURCE_CONFIGS:
        return jsonify({'status': 'error', 'message': 'Invalid sheet ID'}), 400

    config = METALS_SOURCE_CONFIGS[sheet_id]
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'status': 'error'}), 500

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            fmt = config.get('format', 'standard')
            if fmt == 'year_month':
                data = get_brent_data(cur, config)
            elif fmt == 'monthly_matrix':
                data = get_shme_data(cur, config)
            elif fmt == 'yearly_columns':
                data = get_yearly_columns_data(cur, config)
            elif fmt == 'monthly_with_conversion':
                data = get_comex_data(cur, config)
            else:
                data = get_standard_data(cur, config)
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = config['name']

        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="0066B2", end_color="0066B2", fill_type="solid")
        frm_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        ctr      = Alignment(horizontal='center', vertical='center')
        thin     = Side(style='thin', color="CCCCCC")
        bdr      = Border(top=thin, left=thin, right=thin, bottom=thin)

        def style_header(row_cells):
            for c in row_cells:
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = bdr

        if fmt == 'year_month':
            ws.append(['Année', 'Mois', 'Prix Moyen (€/baril)', 'Nb Points'])
            style_header(ws[1])
            for idx, row in enumerate(data, start=2):
                ws[f'A{idx}'] = row.get('year')
                ws[f'B{idx}'] = row.get('month')
                ws[f'C{idx}'] = round(float(row['price']), 2) if row.get('price') else None
                ws[f'D{idx}'] = row.get('data_points')
                for col in 'ABCD':
                    ws[f'{col}{idx}'].border = bdr; ws[f'{col}{idx}'].alignment = ctr

        elif fmt == 'monthly_matrix' and sheet_id == 'shme':
            headers = ['Mois', 'Copper (Non-VAT)', 'Zinc (Non-VAT)', 'Tin (Non-VAT)',
                       'H62', 'H65', 'H68', 'H70', 'H85',
                       'Qsn4-0.1', 'Qsn6.5-0.1', 'Qsn8-0.3', 'T2']
            ws.append(headers); style_header(ws[1])
            for idx, row in enumerate(data, start=2):
                ws[f'A{idx}'] = f"{row['year']}-{row['month']:02d}"
                ws[f'B{idx}'] = row.get('copper_base', 0)
                ws[f'C{idx}'] = row.get('zinc_base', 0)
                ws[f'D{idx}'] = row.get('tin_base', 0)
                ws[f'E{idx}'] = f'=(B{idx}*0.62*1.05+C{idx}*0.38*1.05+4000/1.13)/1000'
                ws[f'F{idx}'] = f'=(B{idx}*0.65*1.05+C{idx}*0.35*1.05+4000/1.13)/1000'
                ws[f'G{idx}'] = f'=(B{idx}*0.68*1.05+C{idx}*0.32*1.05+4000/1.13)/1000'
                ws[f'H{idx}'] = f'=(B{idx}*0.70*1.05+C{idx}*0.30*1.05+4500/1.13)/1000'
                ws[f'I{idx}'] = f'=(B{idx}*0.85*1.05+C{idx}*0.15*1.05+5800/1.13)/1000'
                ws[f'J{idx}'] = f'=(B{idx}*0.96*1.05+D{idx}*0.04*1.05+3000/1.13)/1000'
                ws[f'K{idx}'] = f'=(B{idx}*0.935*1.05+D{idx}*0.065*1.05+3000/1.13)/1000'
                ws[f'L{idx}'] = f'=(B{idx}*0.92*1.05+D{idx}*0.08*1.05+5750/1.13)/1000'
                ws[f'M{idx}'] = f'=(B{idx}*1.05+5500/1.13)/1000'
                for col in 'ABCDEFGHIJKLM':
                    c = ws[f'{col}{idx}']; c.border = bdr; c.alignment = ctr
                    if col in 'EFGHI': c.fill = frm_fill

        elif fmt == 'yearly_columns':
            years      = data.get('years', [])
            rows       = data.get('data', [])
            month_names = ['Jan','Fév','Mar','Avr','Mai','Juin','Juil','Août','Sep','Oct','Nov','Déc']
            ws.append(['Mois'] + [str(y) for y in years])
            style_header(ws[1])
            for m in range(1, 13):
                row_d   = next((r for r in rows if r.get('month') == m), None)
                row_out = [month_names[m-1]]
                for yr in years:
                    val = row_d.get(f'year_{yr}') if row_d else None
                    row_out.append(round(float(val), 2) if val is not None else None)
                ws.append(row_out)
                for col_idx in range(1, len(years)+2):
                    ws.cell(row=m+1, column=col_idx).border = bdr

        elif fmt == 'monthly_with_conversion':
            ws.append(['Année', 'Mois', 'Prix USD/lb', 'Prix USD/kg (×2.203)', 'Formule'])
            style_header(ws[1])
            for idx, row in enumerate(data, start=2):
                ws[f'A{idx}'] = row.get('year')
                ws[f'B{idx}'] = row.get('month')
                ws[f'C{idx}'] = round(float(row['price_lb']), 4) if row.get('price_lb') else None
                ws[f'D{idx}'] = f'=C{idx}*2.203'
                ws[f'E{idx}'] = 'price_lb * 2.203'
                for col in 'ABCDE':
                    c = ws[f'{col}{idx}']; c.border = bdr; c.alignment = ctr
                ws[f'D{idx}'].fill = frm_fill

        else:
            ws.append(['Date', 'Métal', 'Prix', 'Devise', 'Unité'])
            style_header(ws[1])
            for idx, row in enumerate(data, start=2):
                ws[f'A{idx}'] = row.get('price_date')
                ws[f'B{idx}'] = row.get('metal_type')
                ws[f'C{idx}'] = round(float(row['price']), 4) if row.get('price') else None
                ws[f'D{idx}'] = row.get('currency')
                ws[f'E{idx}'] = row.get('unit')
                for col in 'ABCDE':
                    ws[f'{col}{idx}'].border = bdr; ws[f'{col}{idx}'].alignment = ctr

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 10), 30)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"{config['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Erreur export_sheet_excel [{sheet_id}]: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/export/excel')
def export_excel():
    try:
        days       = request.args.get('days', type=int)
        metal_type = request.args.get('metal_type')
        start_date = request.args.get('start_date')
        end_date   = request.args.get('end_date')
        if metal_type and metal_type.lower() == 'all':
            metal_type = None

        history_data = get_price_history(days, metal_type, start_date, end_date)
        if not history_data:
            return jsonify({'status': 'error', 'message': 'Aucune donnée à exporter'}), 404

        date_set = set()
        for item in history_data:
            pd = item['price_date']
            if isinstance(pd, datetime):
                pd = pd.date()
            date_set.add(pd)
        sorted_dates = sorted(list(date_set))

        pivot_data = {}
        for item in history_data:
            metal    = item['metal_type']
            pd       = item['price_date']
            if isinstance(pd, datetime):
                pd = pd.date()
            price    = item['price']
            currency = item['currency']
            unit     = item['unit']
            if metal not in pivot_data:
                pivot_data[metal] = {'currency': currency, 'unit': unit, 'prices': {}}
            pivot_data[metal]['prices'][pd] = price

        wb = Workbook()
        ws = wb.active
        ws.title = "Historique Prix Métaux"
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="0066B2", end_color="0066B2", fill_type="solid")
        data_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
        ctr       = Alignment(horizontal='center', vertical='center')
        lft       = Alignment(horizontal='left',   vertical='center')
        thin      = Side(style='thin', color="CCCCCC")
        bdr       = Border(top=thin, left=thin, right=thin, bottom=thin)

        headers = ['Produit (Metal)', 'Devise', 'Unité'] + [dt.strftime('%Y-%m-%d') for dt in sorted_dates]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr; cell.border = bdr

        row_idx = 2
        for metal, data in pivot_data.items():
            ws.cell(row=row_idx, column=1, value=metal).alignment = lft
            ws.cell(row=row_idx, column=2, value=data['currency']).alignment = ctr
            ws.cell(row=row_idx, column=3, value=data['unit']).alignment = ctr
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).fill = data_fill
                ws.cell(row=row_idx, column=col).border = bdr
            for col_idx, dt in enumerate(sorted_dates, start=4):
                price = data['prices'].get(dt)
                cell = ws.cell(row=row_idx, column=col_idx)
                if price is not None:
                    cell.value = float(price)
                    cell.number_format = '#,##0.########'
                cell.border = bdr; cell.alignment = ctr
            row_idx += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        for col_idx in range(4, len(sorted_dates) + 4):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Prix_Metaux_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Erreur export Excel: {e}")
        return jsonify({'error': str(e)}), 500

# ──────────────────────────────────────────
# API ECB / FX
# ──────────────────────────────────────────
@app.route('/ecb/rates')
def api_ecb_rates():
    start_date     = request.args.get('start_date')
    end_date       = request.args.get('end_date')
    quote_currency = request.args.get('quote_currency')
    month          = request.args.get('month')
    rates = get_ecb_rates(start_date=start_date, end_date=end_date,
                          quote_currency=quote_currency, month=month)
    return jsonify({'status': 'success', 'data': [serialize_row(r) for r in rates]})

@app.route('/ecb/rates/export')
def api_ecb_rates_export():
    try:
        start_date     = request.args.get('start_date')
        end_date       = request.args.get('end_date')
        quote_currency = request.args.get('quote_currency')
        month          = request.args.get('month')
        rates = get_ecb_rates(start_date=start_date, end_date=end_date,
                              quote_currency=quote_currency, month=month)
        if not rates:
            return jsonify({'status': 'error', 'message': 'Aucun taux à exporter'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "ECB FX Rates"
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="0066B2", end_color="0066B2", fill_type="solid")
        ctr      = Alignment(horizontal='center', vertical='center')
        thin     = Side(style='thin', color="CCCCCC")
        bdr      = Border(top=thin, left=thin, right=thin, bottom=thin)

        for col_idx, header in enumerate(['Date', 'Base Currency', 'Quote Currency', 'Rate'], start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr; cell.border = bdr

        for row_idx, r in enumerate(rates, start=2):
            ref_date = r['ref_date']
            if isinstance(ref_date, (datetime, date)):
                ref_date = ref_date if isinstance(ref_date, date) else ref_date.date()
            ws.cell(row=row_idx, column=1, value=ref_date)
            ws.cell(row=row_idx, column=2, value=r['base_currency'])
            ws.cell(row=row_idx, column=3, value=r['quote_currency'])
            rate_cell = ws.cell(row=row_idx, column=4)
            if r['rate'] is not None:
                rate_cell.value = float(r['rate'])
                rate_cell.number_format = '#,##0.0000'
            for col_idx in range(1, 5):
                ws.cell(row=row_idx, column=col_idx).alignment = ctr
                ws.cell(row=row_idx, column=col_idx).border = bdr

        for col, w in zip('ABCD', [12, 14, 16, 14]):
            ws.column_dimensions[col].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"ECB_Rates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Erreur export FX Excel: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/ecb/export-florent')
def export_florent():
    try:
        month = request.args.get('month', type=int)
        year  = request.args.get('year',  type=int)
        if not month or not year:
            return jsonify({'status': 'error', 'message': 'Paramètres month et year requis'}), 400

        month_name = datetime(year, month, 1).strftime('%B %Y')
        data = get_florent_report_data(year, month)
        if not data:
            return jsonify({'status': 'error', 'message': 'Aucune donnée disponible'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly FX Report"
        hdr_fill  = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        hdr_font  = Font(color="FFFFFF", bold=True, size=12)
        bdr_style = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        ctr = Alignment(horizontal='center', vertical='center')
        rgt = Alignment(horizontal='right',  vertical='center')

        headers = [
            'Currency',
            f'Closing Rate ({month_name})',
            'Period Rate (M-1)',
            f'Average YTD ({month_name})',
            'Budget Rate'
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr; cell.border = bdr_style

        for row_data in data:
            ws.append([
                row_data.get('quote_currency'),
                row_data.get('closing_rate'),
                row_data.get('period_rate'),
                row_data.get('ytd_average'),
                row_data.get('budget_rate'),
            ])

        for row in ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row):
                cell.border = bdr_style
                if idx == 0:
                    cell.alignment = ctr
                elif isinstance(cell.value, (float, int)):
                    cell.number_format = '0.0000'
                    cell.alignment = rgt
                elif cell.value is None:
                    cell.value = 'N/A'
                    cell.alignment = ctr

        for col, w in zip('ABCDE', [15, 22, 22, 22, 18]):
            ws.column_dimensions[col].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"AVO_Monthly_FX_{month:02d}_{year}.xlsx"
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Erreur export Florent: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/ecb/monthly-summary')
def api_monthly_fx_summary():
    year           = request.args.get('year',           type=int)
    month          = request.args.get('month',          type=int)
    quote_currency = request.args.get('quote_currency')
    summary = get_monthly_fx_summary(year, month, quote_currency)
    data = [serialize_row(r) for r in summary]
    meta_year  = year  or datetime.now().year
    meta_month = month or datetime.now().month
    metadata = {
        'year':             meta_year,
        'month':            meta_month,
        'month_name':       datetime(meta_year, meta_month, 1).strftime('%B %Y'),
        'is_current_month': (meta_year == datetime.now().year and meta_month == datetime.now().month)
    }
    return jsonify({'status': 'success', 'data': data, 'metadata': metadata})

@app.route('/api/metals/summary')
def api_metals_summary():
    months_param = request.args.get('months', type=int, default=12)
    if months_param < 1 or months_param > 36:
        months_param = 12

    conn = get_db_connection()
    if not conn:
        return jsonify({'status': 'error', 'message': 'DB connection failed'}), 500

    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            today = datetime.now().date()
            periods = []
            y, m = today.year, today.month
            for _ in range(months_param):
                m -= 1
                if m == 0:
                    m = 12; y -= 1
                periods.append(f"{y}-{m:02d}")
            periods.sort(reverse=True)

            # FX USD/EUR
            cur.execute("""
                SELECT TO_CHAR(DATE_TRUNC('month', ref_date), 'YYYY-MM') AS period,
                       AVG(rate) AS fx_rate
                FROM ecb_exchange_rates
                WHERE quote_currency = 'USD'
                  AND ref_date >= %s
                GROUP BY DATE_TRUNC('month', ref_date)
                ORDER BY period DESC
            """, (date(today.year - 3, 1, 1),))
            fx_map = {r['period']: float(r['fx_rate']) for r in cur.fetchall()}

            result_rows = []
            result_rows.append({
                'market': 'FX', 'label': 'USD/EUR', 'metric': 'fx_usd_eur',
                'currency': 'rate', 'decimals': 4,
                'values': {p: fx_map.get(p) for p in periods}
            })

            lme_patterns = ['metals.dev', 'Metal.dev API']
            lme_clauses  = ' OR '.join(["(source_url ILIKE %s OR source_url = %s)"] * len(lme_patterns))
            lme_params   = []
            for p in lme_patterns:
                lme_params.extend([f'%{p}%', p])

            for metal, label_usd, label_eur in [
                ('copper', 'Cu USD/kg', 'Cu €/kg'),
                ('zinc',   'Zn USD/kg', 'Zn €/kg'),
                ('tin',    'Sn USD/kg', 'Sn €/kg'),
            ]:
                cur.execute(f"""
                    SELECT TO_CHAR(DATE_TRUNC('month', price_date), 'YYYY-MM') AS period,
                           AVG(price)/1000 AS avg_price
                    FROM metal_prices
                    WHERE ({lme_clauses}) AND metal_type = %s AND price_date >= %s
                    GROUP BY DATE_TRUNC('month', price_date)
                    ORDER BY period DESC
                """, lme_params + [metal, date(today.year - 3, 1, 1)])
                usd_map = {r['period']: float(r['avg_price']) for r in cur.fetchall()}
                result_rows.append({
                    'market': 'LME', 'label': label_usd, 'metric': f'lme_{metal}_usd',
                    'currency': 'USD', 'decimals': 4,
                    'values': {p: usd_map.get(p) for p in periods}
                })
                result_rows.append({
                    'market': 'LME', 'label': label_eur, 'metric': f'lme_{metal}_eur',
                    'currency': 'EUR', 'decimals': 4,
                    'values': {p: (usd_map[p]/fx_map[p] if p in usd_map and p in fx_map and fx_map[p] else None)
                               for p in periods}
                })

            cu_usd_map = next((r['values'] for r in result_rows if r.get('metric') == 'lme_copper_usd'), {})
            sorted_periods = sorted(periods)
            var_vals = {}
            for i, p in enumerate(sorted_periods[1:], 1):
                prev = sorted_periods[i-1]
                if cu_usd_map.get(p) and cu_usd_map.get(prev):
                    var_vals[p] = (cu_usd_map[p] - cu_usd_map[prev]) / cu_usd_map[prev]
            result_rows.append({
                'market': 'LME', 'label': 'Var Cu USD Δ%', 'metric': 'lme_cu_var',
                'currency': 'USD', 'decimals': 4, 'values': var_vals
            })

            zn_lme = next((r['values'] for r in result_rows if r.get('metric') == 'lme_zinc_usd'), {})
            for alloy, cu_pct, zn_pct in [('CuZn30', 0.70, 0.30), ('CuZn33', 0.67, 0.33), ('CuZn36', 0.64, 0.36)]:
                result_rows.append({
                    'market': 'LME', 'label': f'{alloy} USD/kg', 'metric': f'lme_{alloy.lower()}_usd',
                    'currency': 'USD', 'decimals': 4,
                    'values': {p: (cu_usd_map[p]*cu_pct + zn_lme[p]*zn_pct
                                   if p in cu_usd_map and p in zn_lme else None)
                               for p in periods}
                })

            # COMEX
            cur.execute("""
                SELECT TO_CHAR(DATE_TRUNC('month', price_date), 'YYYY-MM') AS period,
                       AVG(price) * 2.203 AS price_kg_usd
                FROM metal_prices
                WHERE source_url ILIKE '%comexlive%' AND price_date >= %s
                GROUP BY DATE_TRUNC('month', price_date) ORDER BY period DESC
            """, (date(today.year - 3, 1, 1),))
            comex_map = {r['period']: float(r['price_kg_usd']) for r in cur.fetchall()}
            result_rows.append({
                'market': 'COMEX', 'label': 'Cu USD/kg', 'metric': 'comex_cu_usd',
                'currency': 'USD', 'decimals': 4,
                'values': {p: comex_map.get(p) for p in periods}
            })
            result_rows.append({
                'market': 'COMEX', 'label': 'Cu €/kg', 'metric': 'comex_cu_eur',
                'currency': 'EUR', 'decimals': 4,
                'values': {p: (comex_map[p]/fx_map[p] if p in comex_map and p in fx_map and fx_map[p] else None)
                           for p in periods}
            })

            # GIRM
            cur.execute("""
                SELECT TO_CHAR(DATE_TRUNC('month', price_date), 'YYYY-MM') AS period,
                       AVG(price) AS avg_price
                FROM metal_prices
                WHERE source_url ILIKE '%m-lego%' AND price_date >= %s
                GROUP BY DATE_TRUNC('month', price_date) ORDER BY period DESC
            """, (date(today.year - 3, 1, 1),))
            girm_map = {r['period']: (float(r['avg_price'])/100 if float(r['avg_price']) > 30 else float(r['avg_price']))
                        for r in cur.fetchall()}
            result_rows.append({
                'market': 'GIRM', 'label': 'Cu €/kg', 'metric': 'girm_cu_eur',
                'currency': 'EUR', 'decimals': 4,
                'values': {p: girm_map.get(p) for p in periods}
            })

            # LS NIKKO
            cur.execute("""
                SELECT TO_CHAR(DATE_TRUNC('month', price_date), 'YYYY-MM') AS period,
                       AVG(price) AS avg_price
                FROM metal_prices WHERE source_product_name = 'LS Nikko' AND price_date >= %s
                GROUP BY DATE_TRUNC('month', price_date) ORDER BY period DESC
            """, (date(today.year - 3, 1, 1),))
            lsn_map = {r['period']: float(r['avg_price']) for r in cur.fetchall()}
            cur.execute("""
                SELECT TO_CHAR(DATE_TRUNC('month', ref_date), 'YYYY-MM') AS period,
                       AVG(rate) AS fx_rate
                FROM ecb_exchange_rates WHERE quote_currency = 'KRW' AND ref_date >= %s
                GROUP BY DATE_TRUNC('month', ref_date)
            """, (date(today.year - 3, 1, 1),))
            krw_map = {r['period']: float(r['fx_rate']) for r in cur.fetchall()}
            result_rows.append({
                'market': 'LS NIKKO', 'label': 'Cu €/kg', 'metric': 'lsnikko_cu_eur',
                'currency': 'EUR', 'decimals': 4,
                'values': {p: (lsn_map[p] / krw_map[p] / 1000
                               if p in lsn_map and p in krw_map and krw_map[p] else None)
                           for p in periods}
            })

            # SHME
            cur.execute("""
                SELECT TO_CHAR(DATE_TRUNC('month', price_date), 'YYYY-MM') AS period,
                       AVG(price) / 1.13 / 1000 AS cu_nonvat_kg
                FROM metal_prices
                WHERE source_url ILIKE '%shmet%' AND metal_type = 'copper' AND price_date >= %s
                GROUP BY DATE_TRUNC('month', price_date) ORDER BY period DESC
            """, (date(today.year - 3, 1, 1),))
            shme_map = {r['period']: float(r['cu_nonvat_kg']) for r in cur.fetchall()}
            cur.execute("""
                SELECT TO_CHAR(DATE_TRUNC('month', ref_date), 'YYYY-MM') AS period,
                       AVG(rate) AS fx_rate
                FROM ecb_exchange_rates WHERE quote_currency = 'CNY' AND ref_date >= %s
                GROUP BY DATE_TRUNC('month', ref_date)
            """, (date(today.year - 3, 1, 1),))
            cny_map = {r['period']: float(r['fx_rate']) for r in cur.fetchall()}
            result_rows.append({
                'market': 'SHME', 'label': 'Cu CNY/kg (Non-VAT)', 'metric': 'shme_cu_cny',
                'currency': 'CNY', 'decimals': 3,
                'values': {p: shme_map.get(p) for p in periods}
            })
            result_rows.append({
                'market': 'SHME', 'label': 'Cu €/kg', 'metric': 'shme_cu_eur',
                'currency': 'EUR', 'decimals': 4,
                'values': {p: (shme_map[p] / cny_map[p]
                               if p in shme_map and p in cny_map and cny_map[p] else None)
                           for p in periods}
            })

            return jsonify({
                'status': 'success',
                'periods': periods,
                'data': result_rows,
                'metadata': {
                    'months':       months_param,
                    'generated_at': datetime.now().isoformat()
                }
            })
    except Exception as e:
        logger.error(f"Erreur api_metals_summary: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

# ===============================
# POINT D'ENTRÉE
# ===============================
if __name__ == '__main__':
    app.run(debug=True, port=5000)
